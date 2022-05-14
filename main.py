import shutil
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import Progressbar
from openpyxl import load_workbook
import os
import tika
tika.initVM()
from tika import parser
import pyperclip
import math

wb = load_workbook("EXCEL.xlsx")

descs = []

def filesget():
    file = filedialog.askopenfilenames(initialdir="C:\\Downloads", defaultextension=".pdf", filetypes=[("PDF files", ".pdf")])
    global files
    files = file

def submit():

    button1.config(state=DISABLED)
    button2.config(state=DISABLED)

    progressbar.pack(pady=30, padx=30)
    global completedtasks, tasks
    tasks = len(files)
    taskvalue = 100 / tasks
    completedtasks = 0
    taskcounter.set(str(completedtasks) + "/" +str(tasks)+ " tasks completed")
    taskcount.pack()

    def progress():
        global completedtasks, tasks
        progressbar['value'] += taskvalue
        completedtasks += 1
        taskcounter.set(str(completedtasks) + " out of " + str(tasks) + " tasks completed")
        window.update_idletasks()


    for file in files:
        raw = parser.from_file(file)
        global filedata
        filedata = (raw['content'].split(" "))

        def indexing():
            global FreeCells
            FreeCells = []
            for cell in ws["e"]:
                if cell.value == None:
                    FreeCells.append(cell.row)

            startdatecell = "E" + str(FreeCells[0])
            endingdatecell = "F" + str(FreeCells[0])
            readingcell = "G" + str(FreeCells[0])
            dayscount = "I" + str(FreeCells[0])
            addadfee = "L" + str(FreeCells[0])

            ws[startdatecell].value = filedata[filedata.index('Class') + 2][7:17]
            ws[endingdatecell].value = filedata[filedata.index('Class') + 3][5:15]
            ws[readingcell].value = filedata[filedata.index('Capacity:') + 5][:-16]
            ws[dayscount].value = filedata[filedata.index('Class') + 5][6:8]
            ws[addadfee].value = filedata[filedata.index('10.00')]


        date = filedata[filedata.index('Class') + 2][7:17] + " الى " + filedata[filedata.index('Class') + 3][5:15]
        accountdesc = filedata[filedata.index('Summary') + 3]
        readingdesc = " بقراءة " + filedata[filedata.index('Capacity:') + 5][:-16] + " "
        averagecons = int(filedata[filedata.index('Quantity:') + 1][:-13]) / int(filedata[filedata.index('Class') + 5][6:8])
        average = "ومعدل استهلاك " + str(int(math.ceil(averagecons))) + " "
        extratext = "في اليوم من تاريخ "

        if filedata[filedata.index('Summary') + 3] == '10054716660':
            path = "D:\\Downloads\\42"
            filename = filedata[filedata.index('Class') + 3][5:12]
            filename = filename.replace(filename[filename.index("/")], "-") + ".pdf"

            try:
                destination = os.path.join(path, filename)
                shutil.move(file, destination)
            except shutil.SameFileError:
                pass

            ws = wb['2923-42']

            indexing()

            sanadDesc = "سداد شقة 42 حساب رقم "
            finalDesc = sanadDesc + accountdesc + readingdesc + average + extratext + date

            descs.append(finalDesc)

            progress()

        elif filedata[filedata.index('Summary') + 3] == '10062328912':
            path = "D:\\Downloads\\5"
            filename = filedata[filedata.index('Class') + 3][5:12]
            filename = filename.replace(filename[filename.index("/")], "-") + ".pdf"

            try:
                destination = os.path.join(path, filename)
                shutil.move(file, destination)
            except shutil.SameFileError:
                pass

            ws = wb['5675- 5B']

            indexing()

            sanadDesc = "سداد شقة 5 حساب رقم "
            finalDesc = sanadDesc + accountdesc + readingdesc + average + extratext + date

            descs.append(finalDesc)

            progress()

        elif filedata[filedata.index('Summary') + 3] == '30020596395':
            path = "D:\\Downloads\\11"
            filename = filedata[filedata.index('Class') + 3][5:12]
            filename = filename.replace(filename[filename.index("/")], "-") + ".pdf"

            try:
                destination = os.path.join(path, filename)
                shutil.move(file, destination)
            except shutil.SameFileError:
                pass

            ws = wb['11B']

            indexing()

            sanadDesc = "سداد شقة 11 حساب رقم "
            finalDesc = sanadDesc + accountdesc + readingdesc + average + extratext + date

            descs.append(finalDesc)

            progress()

        elif filedata[filedata.index('Summary') + 3] == '30013775583':
            path = "D:\\Downloads\\27"
            filename = filedata[filedata.index('Class') + 3][5:12]
            filename = filename.replace(filename[filename.index("/")], "-") + ".pdf"

            try:
                destination = os.path.join(path, filename)
                shutil.move(file, destination)
            except shutil.SameFileError:
                pass

            ws = wb['27']

            indexing()

            sanadDesc = "سداد شقة 27 حساب رقم "
            finalDesc = sanadDesc + accountdesc + readingdesc + average + extratext + date

            descs.append(finalDesc)

            progress()

        elif filedata[filedata.index('Summary') + 3] == '10054624906':
            path = "D:\\Downloads\\34"
            filename = filedata[filedata.index('Class') + 3][5:12]
            filename = filename.replace(filename[filename.index("/")], "-") + ".pdf"

            try:
                destination = os.path.join(path, filename)
                shutil.move(file, destination)
            except shutil.SameFileError:
                pass

            ws = wb['2932-34']

            indexing()

            sanadDesc = "سداد شقة 34 حساب رقم "
            finalDesc = sanadDesc + accountdesc + readingdesc + average + extratext + date

            descs.append(finalDesc)

            progress()

        elif filedata[filedata.index('Summary') + 3] == '10054717729':
            path = "D:\\Downloads\\44"
            filename = filedata[filedata.index('Class') + 3][5:12]
            filename = filename.replace(filename[filename.index("/")], "-") + ".pdf"

            try:
                destination = os.path.join(path, filename)
                shutil.move(file, destination)
            except shutil.SameFileError:
                pass

            ws = wb['44']

            indexing()

            sanadDesc = "سداد شقة 44 حساب رقم "
            finalDesc = sanadDesc + accountdesc + readingdesc + average + extratext + date

            descs.append(finalDesc)

            progress()

        else:
            pass

    wb.save("EXCEL.xlsx")

    frame2 = Frame(window).pack()
    data_string3 = StringVar()
    data_string3.set(descs)
    listbox = Listbox(frame2, listvariable=data_string3, font = ("Arial", 10), bd= 0, width= 100, fg= "black", borderwidth=6, relief="solid", justify= CENTER)
    listbox.pack(pady=10, padx= 10)

    def printt():
        currentselect = listbox.get(ANCHOR)
        pyperclip.copy(currentselect)

    button3 = Button(frame2, command=printt, text="Copy", font = ("Arial", 20), bd= 0, fg= "black", background= "white").pack()










window = Tk()
window.title("كهرباء شركة المنازل")
image = PhotoImage(file = "Logo.png")
window.iconphoto(True, image)
window.config(background= "#6A172B")
window.geometry("700x700")

frameMain = Frame(window).pack()

data_string = StringVar()
data_string2 = StringVar()
data_string.set("30011715675 - 30020596395 - 30013775583")
entry = Entry(frameMain, state= "readonly", textvariable= data_string, width= 37, font = ("Arial", 20), bd= 0, fg= "black", background= "white").pack()
data_string2.set("30010522932 - 30010522923 - 30029501703")
entry2 = Entry(frameMain, state= "readonly", textvariable= data_string2, width= 37, font = ("Arial", 20), bd= 0, fg= "black", background= "white").pack()

button1 = Button(frameMain, text= "Attach files", background= "white", font= ("Arial", 20), command= filesget)
button1.pack(pady= 20)
button2 = Button(frameMain, text= "Submit", background= "white", font= ("Arial", 20), command= submit)
button2.pack()

progressbar = Progressbar(window, orient=HORIZONTAL, length=300)
taskcounter = StringVar()
taskcount= Label(window, textvariable= taskcounter, background= "white", font = ("Arial", 15))






window.mainloop()
